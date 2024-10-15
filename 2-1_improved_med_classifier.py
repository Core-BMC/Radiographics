import os
import json
from openpyxl import Workbook, load_workbook
from openai import OpenAI

# Initialize OpenAI client
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

if not client.api_key:
    raise ValueError("OPENAI_API_KEY environment variable is not set")

def classify_content_with_gpt4(content):
    """Classify content using GPT-4 model."""
    prompt = f"""
    Carefully analyze the following text and classify it into 6 categories. 
    It is crucial to preserve the original content exactly as it appears, without any modifications.
    Identify the specific sections corresponding to each category and extract them precisely.
    
    Output the result in JSON format with keys '1' to '6'.
    For category 6, include sub-keys '6.1', '6.2', and '6.3'.
    If a category is not present in the original text, leave its value as an empty string.

    The categories are:
    1. Type of Medical Imaging
    2. Specific Imaging Sequence
    3. Use of Contrast
    4. Image Plane
    5. Part of the Body Imaged
    6. Proposing Disease Candidates
       6.1. Names of Three Possible Disease Candidates
       6.2. Likelihood Score for Each Candidate
       6.3. Detailed Rationale for Each Disease

    Ensure that each extracted section maintains its original formatting, including line breaks and punctuation.
    Do not summarize or paraphrase the content; use the exact text from the original document.

    Text to classify:
    {content}

    JSON Output:
    """

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are a precise medical text analyzer that extracts and classifies specific sections from documents without altering the original content."},
                {"role": "user", "content": prompt}
            ],
            temperature=0,
            max_tokens=1000,
            top_p=1,
            frequency_penalty=0,
            presence_penalty=0,
            response_format={"type": "json_object"}
        )

        return json.loads(response.choices[0].message.content)

    except Exception as e:
        print(f"Error processing content: {e}")
        if 'response' in locals():
            print(f"Response content: {response.choices[0].message.content}")
        return {}

def process_file(file_path, sheet):
    """Process a single file and add its content to the Excel sheet."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            classified_content = classify_content_with_gpt4(content)

            sections = []
            for j in range(1, 7):
                section = classified_content.get(str(j), "")
                if isinstance(section, dict):
                    section = "\n".join([f"{k}: {v}" for k, v in section.items()])
                sections.append(section)

            file_number = int(os.path.basename(file_path).split('.')[0])
            sheet.append([file_number] + sections)

            print(f"\nProcessing file: {file_path}")
            for j, section in enumerate(sections, 1):
                print(f"Section {j}:")
                print(section[:100] + "..." if len(section) > 100 else section)
                print("-" * 50)

            empty_items = [idx + 1 for idx, item in enumerate(sections) if not item.strip()]
            return (file_number, empty_items) if empty_items else None

    except FileNotFoundError:
        print(f'File not found: {file_path}')
        return None

def process_folder(folder_path):
    """Process all files in a folder and create a summary Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.append(['Number', '1.', '2.', '3.', '4.', '5.', '6.'])

    problematic_files = []

    for i in range(1, 402):
        file_path = os.path.join(folder_path, f'{i}.png.txt')
        result = process_file(file_path, ws)
        if result:
            problematic_files.append(result)

    excel_file_path = os.path.join(folder_path, 'sum.xlsx')
    wb.save(excel_file_path)
    print(f'Excel file saved: {excel_file_path}')

    return problematic_files

def combine_excel_files(folder_paths):
    """Combine all summary Excel files into a single file."""
    combined_wb = Workbook()
    combined_wb.remove(combined_wb.active)

    for folder_path in folder_paths:
        sum_file_path = os.path.join(folder_path, 'sum.xlsx')
        try:
            wb = load_workbook(sum_file_path)
            ws = wb.active
            sheet_name = folder_path.split('/')[-1]
            new_ws = combined_wb.create_sheet(title=sheet_name)
            for row in ws:
                new_ws.append([cell.value for cell in row])
        except FileNotFoundError:
            print(f'File not found: {sum_file_path}')
            continue

    combined_excel_file_path = 'combined_sum.xlsx'
    combined_wb.save(combined_excel_file_path)
    print(f'Combined Excel file saved: {combined_excel_file_path}')

def main():
    folder_paths = [
        'gpt4o_result/gpt4o_result_temp_0_try1',
        'gpt4o_result/gpt4o_result_temp_0_5_try1',
        'gpt4o_result/gpt4o_result_temp_1_try1',
        'Claude_3.5_result/Claude_3.5_result_temp_0_try1',
        'Claude_3.5_result/Claude_3.5_result_temp_0_5_try1',
        'Claude_3.5_result/Claude_3.5_result_temp_1_try1',
    ]

    all_problematic_files = {}

    for folder_path in folder_paths:
        print(f"Processing folder: {folder_path}")
        problematic_files = process_folder(folder_path)
        if problematic_files:
            all_problematic_files[folder_path] = problematic_files

    combine_excel_files(folder_paths)

    if all_problematic_files:
        print("\nFiles with classification issues:")
        for folder, files in all_problematic_files.items():
            print(f"\n{folder}:")
            for file_num, empty_items in files:
                print(f"  File {file_num}.png.txt: Items {', '.join(map(str, empty_items))} are empty.")
    else:
        print("\nAll files classified successfully.")

if __name__ == "__main__":
    main()