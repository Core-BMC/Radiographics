import os
import pandas as pd
from openpyxl import Workbook, load_workbook

def process_folder(folder_path):
    """Process a single folder and create a summary Excel file."""
    wb = Workbook()
    ws = wb.active

    # Set headers
    ws.append(['Number', '1.', '2.', '3.', '4.', '5.', '6.'])

    for i in range(1, 402):
        file_path = os.path.join(folder_path, f'{i}.png.txt')
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.readlines()

                items = [''] * 6
                current_item_index = 0

                for line in content:
                    stripped_line = line.strip()
                    if current_item_index < 5 and f'{current_item_index + 2}.' in stripped_line:
                        current_item_index += 1
                    items[current_item_index] += line

                ws.append([i] + items)
        except FileNotFoundError:
            print(f'File not found: {file_path}')
            continue

    # Save Excel file
    excel_file_path = os.path.join(folder_path, 'sum.xlsx')
    wb.save(excel_file_path)
    print(f'Excel file saved: {excel_file_path}')

def combine_excel_files(folder_paths):
    """Combine all sum.xlsx files into a single Excel file."""
    combined_wb = Workbook()
    combined_wb.remove(combined_wb.active)  # Remove initial sheet

    for folder_path in folder_paths:
        sum_file_path = os.path.join(folder_path, 'sum.xlsx')
        try:
            wb = load_workbook(sum_file_path)
            ws = wb.active

            sheet_name = folder_path.split('/')[-1]
            new_ws = combined_wb.create_sheet(title=sheet_name)

            for row in ws:
                new_ws.append([cell.value for cell in row])

        except FileNotFoundError:
            print(f'File not found: {sum_file_path}')
            continue

    combined_excel_file_path = 'combined_sum.xlsx'
    combined_wb.save(combined_excel_file_path)
    print(f'Combined Excel file saved: {combined_excel_file_path}')

def main():
    folder_paths = [
        'gpt4o_result/gpt4o_result_temp_0_try1',
        'gpt4o_result/gpt4o_result_temp_0_5_try1',
        'gpt4o_result/gpt4o_result_temp_1_try1',
        'Claude_3.5_result/Claude_3.5_result_temp_0_try1',
        'Claude_3.5_result/Claude_3.5_result_temp_0_5_try1',
        'Claude_3.5_result/Claude_3.5_result_temp_1_try1',
    ]

    # Process each folder
    for folder_path in folder_paths:
        process_folder(folder_path)

    # Combine all Excel files
    combine_excel_files(folder_paths)

if __name__ == "__main__":
    main()